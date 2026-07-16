"""Выгрузка данных марафона в Excel (для P&C)."""
from __future__ import annotations

import hashlib
import hmac
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from bot import db, settings
from bot.config import config

_HEAD = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="2F2B39")
_LINK = Font(color="0563C1", underline="single")


def shot_sig(entry_id: int) -> str:
    """HMAC-подпись ссылки на скриншот: постороннему её не подобрать,
    токен бота в URL не участвует. Проверяется в webapp/server.py."""
    return hmac.new(config.bot_token.encode(), f"shot:{entry_id}".encode(),
                    hashlib.sha256).hexdigest()[:20]


def shot_url(entry_id: int) -> str | None:
    base = settings.webapp_url()
    if not base:
        return None
    return f"{base.rstrip('/')}/shot/{entry_id}/{shot_sig(entry_id)}"


def _fmt(dt) -> str:
    return dt.strftime("%d.%m.%Y %H:%M") if dt else ""


def _status_ru(s: str) -> str:
    return {"accepted": "принят", "pending": "на проверке", "rejected": "отклонён"}.get(s, s)


def _sheet(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for c in ws[1]:
        c.font = _HEAD
        c.fill = _HEAD_FILL
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, 1):
        width = max(len(str(h)), *(len(str(r[i - 1])) for r in rows)) if rows else len(h)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 40)


async def build_workbook() -> tuple[bytes, str]:
    """Собирает xlsx: листы «Участники» и «Результаты». Возвращает (bytes, имя)."""
    wb = Workbook()

    parts = await db.export_participants()
    ws1 = wb.active
    ws1.title = "Участники"
    _sheet(ws1,
           ["ФИО", "Username", "Telegram ID", "Команда", "Наш коллектив", "Статус",
            "Баллов", "Серия", "Регистрация", "Подтверждён"],
           [[p["full_name"], f"@{p['username']}" if p["username"] else "",
             p["telegram_id"], p["team_name"] or "", "да" if p["is_asr"] else "нет",
             "дисквалифицирован" if p["disqualified_at"] else
             ("подтверждён" if p["approved_at"] else "ожидает"),
             p["total_points"], p["streak"], _fmt(p["created_at"]), _fmt(p["approved_at"])]
            for p in parts])

    entries = await db.export_entries()
    ws2 = wb.create_sheet("Результаты")
    _sheet(ws2,
           ["Дата", "ФИО", "Команда", "Шаги", "Баллы", "Статус", "Отправлен",
            "Проверен", "Фото"],
           [[e["entry_date"].strftime("%d.%m.%Y"), e["full_name"], e["team_name"] or "",
             e["steps"], e["points"], _status_ru(e["status"]),
             _fmt(e["created_at"]), _fmt(e["reviewed_at"]),
             "открыть" if e["screenshot_file_id"] else ""]
            for e in entries])
    # Колонка «Фото» — кликабельные ссылки на скриншоты участников.
    photo_col = 9
    for row_i, e in enumerate(entries, start=2):
        if e["screenshot_file_id"]:
            url = shot_url(e["id"])
            if url:
                cell = ws2.cell(row=row_i, column=photo_col)
                cell.hyperlink = url
                cell.font = _LINK

    buf = BytesIO()
    wb.save(buf)
    name = f"step_together_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return buf.getvalue(), name
